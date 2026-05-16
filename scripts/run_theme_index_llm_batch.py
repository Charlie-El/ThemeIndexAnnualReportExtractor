import argparse
import json
import os
import re
import time
from copy import copy
from datetime import datetime
from pathlib import Path

from openai import OpenAI
from openpyxl import Workbook, load_workbook


ROOT_DIR = Path.cwd()
DATA_DIR = ROOT_DIR / "data"
OUTPUTS_DIR = ROOT_DIR / "outputs"
DOWNLOADS_DIR = DATA_DIR / "downloads"
HARD_CODE_TOTAL = OUTPUTS_DIR / "主题指数提取总表.xlsx"
OFFICIAL_OUTPUT_DIR = OUTPUTS_DIR / "主题指数大模型正式输出"
COMPANY_XLSX = "主题指数大模型提取结果.xlsx"
COMPANY_JSON = "主题指数大模型提取结果.json"
COMPANY_RAW = "主题指数大模型原始响应.txt"
EVIDENCE_JSON = "大模型证据包.json"
PROGRESS_JSONL = OFFICIAL_OUTPUT_DIR / "运行明细.jsonl"
TOTAL_XLSX = OFFICIAL_OUTPUT_DIR / "主题指数大模型总表.xlsx"
REPORT_XLSX = OFFICIAL_OUTPUT_DIR / "主题指数大模型运行报告.xlsx"

GOOD_STATUS = "已提取候选内容"
MODEL_NAME = "doubao-seed-1-6-250615"
BASE_URL = "https://ark.cn-beijing.volces.com/api/v3"
MAX_OUTPUT_TOKENS = 16000

TABLE1_FIELDS = [
    "CUSIP",
    "stock_code",
    "company_name",
    "report_period",
    "report_type",
    "industry_code_l1",
    "industry_code_l2",
    "main_business_desc",
    "main_business_desc_summary",
    "total_revenue",
    "segment_disclosure_flag",
    "segment_count",
    "primary_disclosure_dim",
    "disclosure_dims_available",
    "has_revenue_breakdown",
    "has_margin_breakdown",
]

TABLE2_FIELDS = [
    "segment_id",
    "stock_code",
    "report_period",
    "segment_name",
    "segment_type",
    "segment_revenue",
    "segment_revenue_ratio",
    "segment_gross_profit",
    "segment_gross_margin",
    "segment_desc",
    "product_keywords",
    "source_section",
    "mention_sections",
    "is_in_company_intro",
    "is_strategic_focus",
]

EVIDENCE_FIELDS = [
    "field_name",
    "field_value",
    "summary_cn",
    "source_section",
    "source_id",
    "original_text",
    "confidence",
]

TABLE3_FIELDS = [
    "theme_id",
    "theme_name",
    "keyword",
    "keyword_type",
    "keyword_weight",
    "synonym",
    "version",
    "effective_date",
]

TABLE4_FIELDS = [
    "mapping_id",
    "stock_code",
    "report_period",
    "theme_id",
    "relevance_level",
    "relevance_score",
    "theme_revenue_ratio",
    "has_dedicated_segment",
    "keyword_hit_count",
    "keyword_hits",
    "evidence_snippet",
    "evidence_source",
    "confidence",
    "dictionary_version",
    "data_quality_flag",
]

SYSTEM_PROMPT = """你是严谨的美股年报字段抽取员。
你只能根据用户提供的公司映射信息和完整年报证据 JSON 抽取，不得使用外部知识，不得编造。
你的目标是在证据允许的范围内尽量补全字段：能直接引用的直接填，能由已给金额计算的比例可以计算；无法确认的字段必须留空字符串 ""、空数组 [] 或 null。
完整年报证据 JSON 包含 full_text、detected_sections、priority_evidence_blocks、tables 和 relevant_table_order。请优先读 priority_evidence_blocks 和 relevant_table_order 指向的表格，但必要时必须回到 full_text 和完整 tables 查找遗漏。
请只输出合法 JSON，不要输出 Markdown，不要输出解释过程。"""

SCHEMA_PROMPT = """
请严格输出以下 JSON 结构，字段名必须完全一致：
{
  "table1_company_business": {
    "CUSIP": "",
    "stock_code": "",
    "company_name": "",
    "report_period": "",
    "report_type": "",
    "industry_code_l1": "",
    "industry_code_l2": "",
    "main_business_desc": "",
    "main_business_desc_summary": "",
    "total_revenue": "",
    "segment_disclosure_flag": null,
    "segment_count": null,
    "primary_disclosure_dim": "",
    "disclosure_dims_available": [],
    "has_revenue_breakdown": null,
    "has_margin_breakdown": null
  },
  "table2_segment_detail": [
    {
      "segment_id": "",
      "stock_code": "",
      "report_period": "",
      "segment_name": "",
      "segment_type": "",
      "segment_revenue": "",
      "segment_revenue_ratio": "",
      "segment_gross_profit": "",
      "segment_gross_margin": "",
      "segment_desc": "",
      "product_keywords": [],
      "source_section": "",
      "mention_sections": [],
      "is_in_company_intro": null,
      "is_strategic_focus": null
    }
  ],
  "table3_theme_dictionary": [
    {
      "theme_id": "",
      "theme_name": "",
      "keyword": "",
      "keyword_type": "",
      "keyword_weight": "",
      "synonym": [],
      "version": "",
      "effective_date": ""
    }
  ],
  "table4_theme_mapping": [
    {
      "mapping_id": "",
      "stock_code": "",
      "report_period": "",
      "theme_id": "",
      "relevance_level": "",
      "relevance_score": "",
      "theme_revenue_ratio": "",
      "has_dedicated_segment": null,
      "keyword_hit_count": null,
      "keyword_hits": [],
      "evidence_snippet": "",
      "evidence_source": "",
      "confidence": "",
      "dictionary_version": "",
      "data_quality_flag": ""
    }
  ],
  "field_evidence_detail": [
    {
      "field_name": "",
      "field_value": "",
      "summary_cn": "",
      "source_section": "",
      "source_id": "",
      "original_text": "",
      "confidence": ""
    }
  ]
}

字段解释和抽取规则：
1. table1_company_business 对应 PDF 表1“公司主营业务原文”。CUSIP、stock_code、company_name 必须优先照抄公司映射信息。
2. report_period 使用年报覆盖的 fiscal year end，必须输出 YYYY-MM-DD；不要使用下载日期或提交日期。report_type 对 10-K/20-F/40-F 通常输出“年报”。
3. main_business_desc 尽量填写 Item 1 Business / Company Overview 中能代表公司主营业务的原文段落或接近原文的完整句，不要只写模型概括；main_business_desc_summary 用 50-120 字中文概括，尽量包含产品/服务、客户或应用场景。
4. total_revenue、segment_revenue、segment_gross_profit 均用美元整数，不带逗号和单位。若原文写 “$ in millions”，例如 820.627 或 820,627 表示 820627000。
5. segment_revenue_ratio、segment_gross_margin、relevance_score、theme_revenue_ratio、confidence 使用 0-1 小数；如果总收入和分项收入均明确，可以计算占比。
6. primary_disclosure_dim 只能使用 business_segment/product/industry/region/customer_type/narrative。
7. disclosure_dims_available 只能从 business_segment/product/industry/region/customer_type/narrative 中选择。
8. table2_segment_detail 对应 PDF 表2。优先提取 reportable segments / operating segments / business segments；若无业务分部但披露产品、行业、地区、客户收入构成，也要作为表2行提取。若同一公司同时披露 business_segment、product、region、customer_type 多套收入表，优先输出最新年度且最详细的一套，并可补充其他维度，不要只输出一段文字描述。
9. segment_type 只能使用 business_segment/product/industry/region/customer_type/narrative_only。
10. source_section/evidence_source 只能使用 business/MDA/notes/segment/table。
11. 不要把 operating income 填到 segment_gross_profit；如果只有 operating income，没有 gross profit 或 gross margin，则毛利相关字段留空。
12. 禁止输出空对象或空行：如果 table2 某行没有 segment_name 且没有 segment_revenue 且没有 segment_desc，就不要输出这一行。table3/table4 也不要输出空主题行。
13. 如果公司没有披露业务分部或收入分解，table2_segment_detail 输出 []，不要放一行空白占位。
14. 表2只能放主营业务收入、产品收入、服务收入、合作/特许权收入、地区收入、客户收入或经营分部收入。不要把 Research Tax Credits/CIR、Subsidies、Depositary service fees、interest income、investment income、grant income、tax benefit、other income、finance income、foreign exchange gain/loss 填成业务分部或收入构成。
15. 临床阶段药企如果明确“尚未产生产品销售收入/不预计近期产生产品收入”，且只披露税收抵免、补贴或存托服务费等其他收入，table2_segment_detail 应输出 []，total_revenue 也不要用这些其他收入冒充主营收入。
16. data_quality_flag 只能填写 A/B/C/D/E 单个字母。A=分部+营收+占比，B=产品/行业/地区/客户+营收，C=有总营收无分解，D=仅文字描述，E=材料不足。
17. 所有能提取的字段都要提取；确实没有证据的字段留空，不要删除字段，不要根据常识补写。
18. 如果表格中同时出现收入、成本、毛利或调整后毛利，请区分：segment_revenue 填收入；segment_gross_profit 只在原文是 gross profit/gross margin/segment adjusted gross margin 且可对应分部时填写；不要把 cost of sales 当成毛利。
19. source_section 优先使用 PDF需求优先证据里的 source，允许值仍为 business/MDA/notes/segment/table。来自高分表格候选但未识别章节时填 table。
20. primary_disclosure_dim 应代表最终采用的最详细收入披露维度，而不是固定优先级。若有业务分部收入表，用 business_segment；若没有业务分部但有产品/行业/地区/客户收入表，优先选择“更能说明业务构成”的产品/行业表，其次地区表，最后客户表；如果地区表行数更多但只说明销售区域，不要让它覆盖产品/行业维度。
21. 如果同时存在 Automotive/Industrial、Magnetic sensors/Power IC、geographic 等收入表，table2 必须优先输出 Automotive/Industrial 或产品线收入，地区表可作为补充但不应作为唯一结果。
22. 如果源文件质量提示说明 40-F 主要是 Principal Documents/Exhibit 引用页，且没有 Exhibit 99.1/99.2/99.3 正文，不要从 forward-looking statements、风险因素或 Exhibit 标题推断主营业务；main_business_desc 可留空或仅说明“材料仅提供 Exhibit 引用，未提供正文”，table2 留空。
23. field_evidence_detail 用于人工复核。请至少覆盖主营业务、分部/业务构成、收入构成、收入确认、主题相关性这些字段；original_text 必须放可复核的英文原文句子、段落或表格行，summary_cn 放中文解释。不要只给中文总结。
24. original_text 可以较长，优先保留完整句子或完整表格行；如一个字段来自表格，请保留表头和相关行，避免只截单个数字。

主题ID候选：
- THEME_AI：AI、人工智能、加速计算、数据中心AI基础设施、模型训练、推理、GPU算力。
- THEME_SEMICONDUCTOR：半导体、芯片、晶圆、EDA、封装测试、存储、模拟、逻辑、传感器、功率IC。
- THEME_ROBOTICS：机器人、自动驾驶、无人系统、机器视觉、运动控制、工业自动化。
- THEME_SPACE：商业航天、卫星、火箭、低轨、航天系统、卫星通信、高空平台。
- THEME_GEOPOLITICAL_RISK：国防、军工、网络安全、制裁、出口管制、国家安全。
- THEME_ENERGY：油气、天然气、LNG、炼油、可再生能源、电力、储能。
- THEME_MINERALS：矿业、黄金、白银、铜、铝、锂、稀土、贵金属。
- THEME_BIOMEDICAL：生物科技、药物、临床试验、疫苗、基因治疗、医疗设备、诊断。

主题输出规则：
1. table3_theme_dictionary 只输出本公司材料中实际命中的主题关键词候选；keyword_weight/version/effective_date 无法从材料确认则留空。
2. table4_theme_mapping 只输出有证据支持的主题。若没有证据，不要为了覆盖主题而输出 none 行。
3. relevance_level 只能用 high/medium/low/none。主营业务高度围绕该主题填 high；重要产品/收入分部相关填 high/medium；仅风险暴露或辅助技术填 low/medium。
4. has_dedicated_segment 表示是否存在专门以该主题命名或明显归属于该主题的收入/业务分部。普通业务分部包含主题应用但无法完全归因时，不要强行填 true。
5. theme_revenue_ratio 只有能根据分部/产品收入明确归因到主题时才填；否则留空。
"""

GENERIC_TARGET_TERMS = [
    "net revenues",
    "net revenue",
    "total revenue",
    "total revenues",
    "revenue",
    "revenues",
    "net sales",
    "sales",
    "reportable segments",
    "operating segments",
    "business segments",
    "single reportable segment",
    "one reportable segment",
    "chief operating decision maker",
    "segment information",
    "segments revenue",
    "revenue by reportable segment",
    "segment net sales",
    "segment revenue",
    "disaggregated revenue",
    "disaggregated revenues",
    "disaggregation of revenue",
    "disaggregation of revenues",
    "revenue by",
    "revenues by",
    "net sales by",
    "sales by",
    "product revenue",
    "product revenues",
    "service revenue",
    "collaboration revenue",
    "royalty revenue",
    "license revenue",
    "contract revenue",
    "geographic information",
    "geographical information",
    "geographic revenue",
    "revenue by geographic",
    "revenues by geographic",
    "international revenue",
    "major customers",
    "customer concentration",
    "revenue recognition",
    "recognition of revenue",
    "net sales are recognized",
    "revenue is recognized",
    "revenues are recognized",
    "transfer of control",
    "contracts with customers",
    "performance obligations",
    "deferred revenue",
    "contract liabilities",
    "business",
    "overview",
    "results of operations",
    "business overview",
    "company overview",
    "description of business",
    "our business",
    "our businesses",
    "our company",
    "products and services",
    "product offerings",
    "principal products",
    "principal services",
    "industry",
    "customers",
    "artificial intelligence",
    "semiconductor",
    "robotics",
    "autonomous",
    "satellite",
    "aerospace",
    "defense",
    "export controls",
    "cybersecurity",
    "oil",
    "natural gas",
    "LNG",
    "renewable",
    "gold",
    "copper",
    "lithium",
    "clinical trial",
    "biotechnology",
    "pharmaceutical",
]


FORMAL_USER_INSTRUCTION = """
正式抽取要求：
1. 这是正式生产结果，不是测试。请严格基于我提供的“完整年报证据 JSON”抽取。
2. 完整年报证据 JSON 是本地程序从公司年报 HTML 中生成的证据包，不是模型总结；其中 full_text、detected_sections、priority_evidence_blocks 和 tables 都要作为可用原文。
3. 工作顺序建议为：先读 source_quality_notes；再读 priority_evidence_blocks；再按 relevant_table_order 检查 tables 中的高分表格；最后回到 detected_sections 和 full_text 补漏。
4. PDF需求优先证据的口径为：P1=Item 8/Notes 的 Segment Information 或收入构成；P2=Item 8/Notes 的 REVENUES；P3=Item 1 Business 或 Item 7 MD&A；P4=Revenue Recognition。
5. 同一字段有多个候选时，优先使用最新 fiscal year 的完整表格；若较低优先级来源明显更详细，可以补充，但必须在 source_section/evidence_snippet 中保留来源。
6. 不能为了减少空缺而伪造。允许的“补全”只有三类：照抄公司映射信息、从明确金额计算占比、把英文原文准确翻译/概括为中文。
7. 表2的分部明细优先使用公司披露的 reportable segments / operating segments；如果没有业务分部，但有产品、行业/应用、地区、客户或合同类型收入构成，也要作为表2行提取。
8. 不要只输出一段文字描述。只要原文表格或完整句子能支持，就应输出结构化行：分部、产品线、行业/应用、地区、客户类型等能提取的维度都要尽量输出。
9. 如果同一公司同时有多张收入构成表，优先输出最能说明业务构成的表；其他维度如果有明确金额也可以补充到 table2，segment_type 写清楚。
10. 如果某个字段找不到证据，留空即可；不要输出一行全空的分部、主题或映射。
11. evidence_snippet 必须来自完整年报证据 JSON 中的原文短句或表格行，不能自己编写。
12. main_business_desc 尽量保留原文完整句或完整段落，不要只写中文概括；中文概括放在 main_business_desc_summary。
13. 如果材料是 40-F 且只包含 Exhibit 索引，没有 Exhibit 正文，请只根据索引和已有正文提取，缺少的收入/分部/附注字段留空；不要把 forward-looking statements 中的矿山、储量、生产计划等风险提示词整理成主营业务。
"""


def clean_json_text(text: str) -> str:
    text = (text or "").strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?", "", text, flags=re.I).strip()
        text = re.sub(r"```$", "", text).strip()
    start = text.find("{")
    end = text.rfind("}")
    if start >= 0 and end > start:
        return text[start : end + 1]
    return text


def safe_cell(value, limit: int = 32000) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        value = json.dumps(value, ensure_ascii=False)
    value = str(value)
    if len(value) <= limit:
        return value
    return value[: limit - 20] + "\n...[已截断]"


def write_table_sheet(wb: Workbook, title: str, rows, fields: list[str]) -> None:
    ws = wb.create_sheet(title)
    ws.append(fields)
    if isinstance(rows, dict):
        rows = [rows]
    for row in rows:
        ws.append([safe_cell(row.get(field, "")) for field in fields])
    for index, _ in enumerate(fields, start=1):
        ws.column_dimensions[chr(64 + index) if index <= 26 else "A"].width = 24


def find_mapping_workbook(search_dir: Path = DATA_DIR) -> Path | None:
    if not search_dir.exists():
        return None
    for path in search_dir.glob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            wb.close()
            if {"CUSIP", "SECU_CODE", "SECUNAME"}.issubset(set(headers)):
                return path
        except Exception:
            continue
    return None


def load_company_mapping(path: Path | None = None) -> dict[str, dict]:
    if path is None:
        path = find_mapping_workbook()
    if not path:
        return {}
    if not path.exists():
        raise FileNotFoundError(f"公司映射文件不存在: {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    headers = list(rows[0])
    idx = {name: headers.index(name) for name in ["CUSIP", "SECU_CODE", "SECUNAME"] if name in headers}
    mapping = {}
    for row in rows[1:]:
        name = str(row[idx["SECUNAME"]] or "").strip()
        if not name:
            continue
        mapping[name.lower()] = {
            "CUSIP": str(row[idx["CUSIP"]] or "").strip(),
            "stock_code": str(row[idx["SECU_CODE"]] or "").strip(),
            "company_name": name,
        }
    return mapping


def normalize_object(obj: dict, fields: list[str]) -> dict:
    bool_or_number = {
        "segment_disclosure_flag",
        "has_revenue_breakdown",
        "has_margin_breakdown",
        "is_in_company_intro",
        "is_strategic_focus",
        "has_dedicated_segment",
        "keyword_hit_count",
        "segment_count",
    }
    return {field: obj.get(field, None if field in bool_or_number else "") for field in fields}


def has_any_value(row: dict, fields: list[str]) -> bool:
    for field in fields:
        value = row.get(field)
        if isinstance(value, list) and value:
            return True
        if value is not None and str(value).strip():
            return True
    return False


NON_OPERATING_SEGMENT_TERMS = [
    "research tax credit",
    "research tax credits",
    "tax credit",
    "tax credits",
    "cir",
    "subsidy",
    "subsidies",
    "depositary service fee",
    "depositary service fees",
    "interest income",
    "investment income",
    "finance income",
    "financial income",
    "other income",
    "foreign exchange",
    "grant income",
    "tax benefit",
]


def is_non_operating_segment(row: dict) -> bool:
    text = " ".join(
        str(row.get(field) or "").lower()
        for field in ["segment_name", "segment_type", "segment_desc", "source_section"]
    )
    return any(term in text for term in NON_OPERATING_SEGMENT_TERMS)


def normalize_list(value) -> list:
    return value if isinstance(value, list) else []


def coerce_list_field(row: dict, field: str) -> None:
    value = row.get(field)
    if value is None or value == "":
        row[field] = []
    elif not isinstance(value, list):
        row[field] = [str(value)]


def normalize_dimension(value) -> str:
    text = str(value or "").strip()
    mapping = {
        "market": "industry",
        "end_market": "industry",
        "end market": "industry",
        "application": "industry",
        "applications": "industry",
        "business": "business_segment",
        "segment": "business_segment",
        "geography": "region",
        "geographic": "region",
        "customer": "customer_type",
    }
    return mapping.get(text, text)


def normalize_result(data: dict, mapping_info: dict | None = None) -> dict:
    table1 = normalize_object(data.get("table1_company_business") or {}, TABLE1_FIELDS)
    mapping_info = mapping_info or {}
    for source, target in [("CUSIP", "CUSIP"), ("stock_code", "stock_code"), ("company_name", "company_name")]:
        if mapping_info.get(source):
            table1[target] = mapping_info[source]
    dims_available = table1.get("disclosure_dims_available")
    if dims_available == "" or dims_available is None:
        table1["disclosure_dims_available"] = []
    elif not isinstance(dims_available, list):
        table1["disclosure_dims_available"] = [str(table1["disclosure_dims_available"])]
    table1["primary_disclosure_dim"] = normalize_dimension(table1.get("primary_disclosure_dim"))
    table1["disclosure_dims_available"] = sorted(
        {
            normalize_dimension(item)
            for item in table1["disclosure_dims_available"]
            if normalize_dimension(item) in {"business_segment", "product", "industry", "region", "customer_type", "narrative"}
        }
    )

    stock_code = table1.get("stock_code") or mapping_info.get("stock_code", "")
    report_period = table1.get("report_period", "")

    table2 = []
    for item in normalize_list(data.get("table2_segment_detail")):
        row = normalize_object(item or {}, TABLE2_FIELDS)
        coerce_list_field(row, "product_keywords")
        coerce_list_field(row, "mention_sections")
        if stock_code and not row.get("stock_code"):
            row["stock_code"] = stock_code
        if report_period and not row.get("report_period"):
            row["report_period"] = report_period
        row["segment_type"] = normalize_dimension(row.get("segment_type"))
        if is_non_operating_segment(row):
            continue
        if has_any_value(row, ["segment_name", "segment_type", "segment_revenue", "segment_desc", "product_keywords"]):
            table2.append(row)

    table3 = []
    for item in normalize_list(data.get("table3_theme_dictionary")):
        row = normalize_object(item or {}, TABLE3_FIELDS)
        coerce_list_field(row, "synonym")
        if has_any_value(row, ["theme_id", "theme_name", "keyword"]):
            table3.append(row)

    table4 = []
    for item in normalize_list(data.get("table4_theme_mapping")):
        row = normalize_object(item or {}, TABLE4_FIELDS)
        coerce_list_field(row, "keyword_hits")
        if stock_code and not row.get("stock_code"):
            row["stock_code"] = stock_code
        if report_period and not row.get("report_period"):
            row["report_period"] = report_period
        if has_any_value(row, ["theme_id", "relevance_level", "keyword_hits", "evidence_snippet"]):
            table4.append(row)

    evidence_detail = []
    for item in normalize_list(data.get("field_evidence_detail")):
        row = normalize_object(item or {}, EVIDENCE_FIELDS)
        if has_any_value(row, ["field_name", "field_value", "summary_cn", "original_text"]):
            evidence_detail.append(row)

    segment_count = table1.get("segment_count")
    if (segment_count == "" or segment_count is None) and table2:
        table1["segment_count"] = len(table2)
    if table1.get("segment_disclosure_flag") is None and table2:
        table1["segment_disclosure_flag"] = any(row.get("segment_type") == "business_segment" for row in table2)
    if table1.get("has_revenue_breakdown") is None and table2:
        table1["has_revenue_breakdown"] = any(str(row.get("segment_revenue") or "").strip() for row in table2)
    if table1.get("has_margin_breakdown") is None and table2:
        table1["has_margin_breakdown"] = any(
            str(row.get("segment_gross_profit") or "").strip() or str(row.get("segment_gross_margin") or "").strip()
            for row in table2
        )

    return {
        "table1_company_business": table1,
        "table2_segment_detail": table2,
        "table3_theme_dictionary": table3,
        "table4_theme_mapping": table4,
        "field_evidence_detail": evidence_detail,
    }


def read_good_company_names(path: Path, status: str = GOOD_STATUS) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["总表"] if "总表" in wb.sheetnames else wb.worksheets[0]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        name_idx = headers.index("公司名称")
        status_idx = headers.index("提取状态")
        names = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[name_idx] or "").strip()
            row_status = str(row[status_idx] or "").strip()
            if name and row_status == status:
                names.append(name)
        return names
    finally:
        wb.close()


def read_compact_model_input(folder: Path) -> str:
    workbook = folder / "主题指数提取结果.xlsx"
    if not workbook.exists():
        return ""
    wb = load_workbook(workbook, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[-1]
        for row in ws.iter_rows(values_only=True):
            if row and str(row[0]) == "建议输入":
                return str(row[1] or "")
    finally:
        wb.close()
    return ""


def read_evidence_package(folder: Path) -> dict:
    path = folder / EVIDENCE_JSON
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8", errors="ignore"))


def evidence_package_as_text(folder: Path) -> str:
    evidence = read_evidence_package(folder)
    if not evidence:
        return ""
    return json.dumps(evidence, ensure_ascii=False, indent=2)


def section_between(text: str, start_marker: str, end_marker: str) -> str:
    start = text.find(start_marker)
    if start < 0:
        return ""
    start += len(start_marker)
    end = text.find(end_marker, start)
    if end < 0:
        end = len(text)
    return text[start:end].strip()


def read_section_input_body(folder: Path) -> str:
    path = folder / "大模型章节输入.txt"
    if not path.exists():
        return ""
    text = path.read_text(encoding="utf-8", errors="ignore")
    blocks = [
        section_between(text, "===== PDF需求优先证据开始 =====", "===== PDF需求优先证据结束 ====="),
        section_between(text, "===== 高分表格候选开始 =====", "===== 高分表格候选结束 ====="),
        section_between(text, "===== 重点表格候选开始 =====", "===== 重点表格候选结束 ====="),
        section_between(text, "===== 重点文本窗口候选开始 =====", "===== 重点文本窗口候选结束 ====="),
        section_between(text, "===== 年报章节开始 =====", "===== 年报章节结束 ====="),
        section_between(text, "===== 原文兜底片段开始 =====", "===== 原文兜底片段结束 ====="),
    ]
    return "\n\n".join(block for block in blocks if block)


def read_named_sections(folder: Path) -> dict[str, str]:
    path = folder / "大模型章节输入.txt"
    if not path.exists():
        return {}
    text = path.read_text(encoding="utf-8", errors="ignore")
    return {
        "priority": section_between(text, "===== PDF需求优先证据开始 =====", "===== PDF需求优先证据结束 ====="),
        "high_score_tables": section_between(text, "===== 高分表格候选开始 =====", "===== 高分表格候选结束 ====="),
        "tables": section_between(text, "===== 重点表格候选开始 =====", "===== 重点表格候选结束 ====="),
        "windows": section_between(text, "===== 重点文本窗口候选开始 =====", "===== 重点文本窗口候选结束 ====="),
        "sections": section_between(text, "===== 年报章节开始 =====", "===== 年报章节结束 ====="),
        "fallbacks": section_between(text, "===== 原文兜底片段开始 =====", "===== 原文兜底片段结束 ====="),
    }


def collect_windows(text: str, terms: list[str], label: str, before: int = 1600, after: int = 5200, limit: int = 6) -> list[str]:
    if not text:
        return []
    lower = text.lower()
    snippets = []
    seen = set()
    for term in terms:
        term_lower = term.lower()
        start = 0
        while True:
            idx = lower.find(term_lower, start)
            if idx < 0:
                break
            left = max(0, idx - before)
            right = min(len(text), idx + after)
            snippet = text[left:right].strip()
            key = re.sub(r"\s+", " ", snippet[:600]).lower()
            if key not in seen:
                snippets.append(f"===== {label}：{term} =====\n{snippet}")
                seen.add(key)
            start = idx + len(term)
            if len(snippets) >= limit:
                return snippets
    return snippets


def extract_targeted_snippets(folder: Path) -> str:
    named = read_named_sections(folder)
    if not any(named.values()):
        return ""

    priority_block = named.get("priority", "")[:26000]
    high_score_tables = named.get("high_score_tables", "")[:18000]
    table_block = named.get("tables", "")[:12000]
    windows_block = named.get("windows", "")[:8000]
    section_block = named.get("sections", "")
    fallback_block = named.get("fallbacks", "")

    business_terms = [
        "our company",
        "our business",
        "our businesses",
        "business overview",
        "company overview",
        "description of business",
        "products and services",
        "product offerings",
        "principal products",
        "principal services",
        "customers",
        "markets",
    ]
    revenue_terms = [
        "net revenues",
        "net revenue",
        "total revenue",
        "revenues by",
        "revenue by",
        "net sales by",
        "sales by",
        "disaggregated revenue",
        "product revenue",
        "service revenue",
        "collaboration revenue",
        "royalty revenue",
        "geographic",
        "major customers",
    ]
    segment_terms = [
        "reportable segments",
        "operating segments",
        "business segments",
        "single reportable segment",
        "chief operating decision maker",
        "segment information",
    ]
    recognition_terms = [
        "revenue recognition",
        "recognition of revenue",
        "revenue is recognized",
        "revenues are recognized",
        "net sales are recognized",
        "transfer of control",
        "performance obligations",
        "deferred revenue",
        "contract liabilities",
    ]
    theme_terms = [
        "artificial intelligence",
        "AI",
        "semiconductor",
        "robotics",
        "autonomous",
        "satellite",
        "aerospace",
        "defense",
        "export controls",
        "cybersecurity",
        "oil",
        "natural gas",
        "LNG",
        "renewable",
        "gold",
        "copper",
        "lithium",
        "clinical trial",
        "biotechnology",
        "pharmaceutical",
    ]

    snippets = []
    if priority_block:
        snippets.append("===== 优先阅读：PDF需求优先证据 =====\n" + priority_block)
    if high_score_tables:
        snippets.append("===== 优先阅读：高分表格候选 =====\n" + high_score_tables)
    if table_block:
        snippets.append("===== 优先阅读：重点表格候选 =====\n" + table_block)
    if windows_block:
        snippets.append("===== 优先阅读：重点文本窗口候选 =====\n" + windows_block)
    snippets.extend(collect_windows(section_block, business_terms, "业务/主营业务原文片段", limit=5))
    snippets.extend(collect_windows(section_block, revenue_terms, "收入构成原文片段", limit=6))
    snippets.extend(collect_windows(section_block, segment_terms, "分部披露原文片段", limit=5))
    snippets.extend(collect_windows(section_block, recognition_terms, "收入确认原文片段", limit=4))
    snippets.extend(collect_windows(section_block + "\n\n" + fallback_block, theme_terms, "主题相关原文片段", before=1400, after=4200, limit=6))

    combined = "\n\n".join(snippets)
    if combined.strip():
        return combined[:52000]
    return read_section_input_body(folder)[:52000]


def build_prompt(company: str, folder: Path, mapping_info: dict) -> str:
    evidence_json = evidence_package_as_text(folder)
    if not evidence_json:
        compact = read_compact_model_input(folder)
        targeted = extract_targeted_snippets(folder)
        evidence_json = json.dumps(
            {
                "warning": f"未找到 {EVIDENCE_JSON}，临时使用旧版候选文本。建议先运行 build_llm_section_inputs.py 生成完整 JSON 证据包。",
                "compact_model_input": compact,
                "targeted_snippets": targeted,
            },
            ensure_ascii=False,
            indent=2,
        )
    payload = [
        FORMAL_USER_INSTRUCTION,
        "\n输出字段口径：",
        SCHEMA_PROMPT,
        "\n公司映射信息：",
        json.dumps(mapping_info, ensure_ascii=False, indent=2),
        "\n完整年报证据 JSON：",
        evidence_json,
    ]
    return "\n".join(payload)


def call_doubao(client: OpenAI, prompt: str) -> tuple[dict, str]:
    completion = client.chat.completions.create(
        model=MODEL_NAME,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": prompt},
        ],
        temperature=0,
    )
    raw = completion.choices[0].message.content
    data = json.loads(clean_json_text(raw))
    return data, raw


def call_doubao_raw(client: OpenAI, prompt: str) -> str:
    completion = client.chat.completions.create(
        model=MODEL_NAME,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": prompt},
        ],
        temperature=0,
        max_tokens=MAX_OUTPUT_TOKENS,
    )
    return completion.choices[0].message.content or ""


def parse_doubao_json(raw: str) -> dict:
    return json.loads(clean_json_text(raw))


def company_done(folder: Path) -> bool:
    return (folder / COMPANY_JSON).exists() and (folder / COMPANY_XLSX).exists()


def write_company_outputs(folder: Path, data: dict, raw: str) -> None:
    xlsx_path = folder / COMPANY_XLSX
    json_path = folder / COMPANY_JSON
    raw_path = folder / COMPANY_RAW

    wb = Workbook()
    wb.remove(wb.active)
    write_table_sheet(wb, "table1_company_business", data["table1_company_business"], TABLE1_FIELDS)
    write_table_sheet(wb, "table2_segment_detail", data["table2_segment_detail"], TABLE2_FIELDS)
    write_table_sheet(wb, "table3_theme_dictionary", data["table3_theme_dictionary"], TABLE3_FIELDS)
    write_table_sheet(wb, "table4_theme_mapping", data["table4_theme_mapping"], TABLE4_FIELDS)
    write_table_sheet(wb, "field_evidence_detail", data.get("field_evidence_detail", []), EVIDENCE_FIELDS)
    info = wb.create_sheet("run_info")
    info.append(["字段", "内容"])
    info.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    info.append(["模型", MODEL_NAME])
    info.append(["说明", "正式大模型抽取结果"])
    info.column_dimensions["A"].width = 18
    info.column_dimensions["B"].width = 80
    wb.save(xlsx_path)
    wb.close()

    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    raw_path.write_text(raw or "", encoding="utf-8")


def load_existing_company_result(folder: Path) -> dict | None:
    path = folder / COMPANY_JSON
    if not path.exists():
        return None
    try:
        return normalize_result(json.loads(path.read_text(encoding="utf-8", errors="ignore")))
    except Exception:
        return None


def append_progress(row: dict) -> None:
    OFFICIAL_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with PROGRESS_JSONL.open("a", encoding="utf-8") as f:
        f.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_final_workbooks(results: list[dict], report_rows: list[dict]) -> None:
    OFFICIAL_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    table1_rows = []
    table2_rows = []
    table3_rows = []
    table4_rows = []
    evidence_rows = []
    warnings = []
    for item in results:
        data = item["data"]
        company = item["company"]
        table1_rows.append(data["table1_company_business"])
        for row in data["table2_segment_detail"]:
            row = dict(row)
            if not row.get("stock_code"):
                row["stock_code"] = data["table1_company_business"].get("stock_code", "")
            table2_rows.append(row)
        for row in data["table3_theme_dictionary"]:
            table3_rows.append(dict(row))
        for row in data["table4_theme_mapping"]:
            row = dict(row)
            if not row.get("stock_code"):
                row["stock_code"] = data["table1_company_business"].get("stock_code", "")
            table4_rows.append(row)
        for row in data.get("field_evidence_detail", []):
            row = dict(row)
            row["company_name"] = data["table1_company_business"].get("company_name", company)
            evidence_rows.append(row)
        if not data["table2_segment_detail"]:
            warnings.append({"company": company, "message": "table2_segment_detail为空"})

    write_table_sheet(wb, "table1_company_business", table1_rows, TABLE1_FIELDS)
    write_table_sheet(wb, "table2_segment_detail", table2_rows, TABLE2_FIELDS)
    write_table_sheet(wb, "table3_theme_dictionary", table3_rows, TABLE3_FIELDS)
    write_table_sheet(wb, "table4_theme_mapping", table4_rows, TABLE4_FIELDS)
    write_table_sheet(wb, "field_evidence_detail", evidence_rows, ["company_name"] + EVIDENCE_FIELDS)
    wb.save(TOTAL_XLSX)
    wb.close()

    report = Workbook()
    ws = report.active
    ws.title = "运行报告"
    ws.append(["项目", "数值"])
    ws.append(["运行时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["成功企业数", sum(1 for row in report_rows if row.get("status") in {"success", "skipped_existing"})])
    ws.append(["失败企业数", sum(1 for row in report_rows if row.get("status") == "error")])
    ws.append(["总表路径", str(TOTAL_XLSX)])

    detail = report.create_sheet("明细")
    fields = ["company", "status", "message", "xlsx", "json", "elapsed_seconds"]
    detail.append(fields)
    for row in report_rows:
        detail.append([safe_cell(row.get(field, "")) for field in fields])

    warning_ws = report.create_sheet("提示")
    warning_ws.append(["company", "message"])
    for row in warnings:
        warning_ws.append([safe_cell(row.get("company", "")), safe_cell(row.get("message", ""))])
    for sheet in report.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                alignment = copy(cell.alignment)
                alignment.wrap_text = True
                alignment.vertical = "top"
                cell.alignment = alignment
        for cell in sheet[1]:
            font = copy(cell.font)
            font.bold = True
            cell.font = font
    report.save(REPORT_XLSX)
    report.close()


def main() -> None:
    global MODEL_NAME
    parser = argparse.ArgumentParser(description="Run official Doubao extraction for good-quality theme-index companies.")
    parser.add_argument("--downloads-dir", default=str(DOWNLOADS_DIR))
    parser.add_argument("--status", default=GOOD_STATUS, help="Only run companies with this status in 主题指数提取总表.xlsx.")
    parser.add_argument("--companies", nargs="*", default=[], help="Optional explicit company names.")
    parser.add_argument("--limit", type=int, default=0, help="Optional max companies after filtering.")
    parser.add_argument("--rerun", action="store_true", help="Rerun even if official company outputs already exist.")
    parser.add_argument("--sleep", type=float, default=0.5, help="Seconds to sleep between model calls.")
    parser.add_argument("--hard-code-total", default=str(HARD_CODE_TOTAL), help="Path to 主题指数提取总表.xlsx generated by theme_index_extractor.py.")
    parser.add_argument("--mapping-file", default="", help="Optional company mapping workbook with CUSIP, SECU_CODE and SECUNAME columns.")
    parser.add_argument("--api-key-env", default="ARK_API_KEY", help="Environment variable that stores the Ark/OpenAI-compatible API key.")
    parser.add_argument("--model-name", default=MODEL_NAME, help="Ark/OpenAI-compatible model or endpoint id.")
    args = parser.parse_args()

    api_key = os.environ.get(args.api_key_env)
    if not api_key:
        raise SystemExit(f"未找到环境变量 {args.api_key_env}，无法调用模型。")

    downloads_dir = Path(args.downloads_dir)
    if args.companies:
        good_names = args.companies
    else:
        good_names = read_good_company_names(Path(args.hard_code_total), args.status)
    if args.limit > 0:
        good_names = good_names[: args.limit]

    folder_map = {folder.name.lower(): folder for folder in downloads_dir.iterdir() if folder.is_dir()}
    mapping_file = Path(args.mapping_file) if args.mapping_file else None
    mapping = load_company_mapping(mapping_file)
    MODEL_NAME = args.model_name
    client = OpenAI(base_url=BASE_URL, api_key=api_key)

    print(f"selected_companies={len(good_names)} status={args.status}", flush=True)
    results = []
    report_rows = []

    for index, company in enumerate(good_names, start=1):
        started = time.time()
        folder = folder_map.get(company.lower())
        if not folder:
            row = {"company": company, "status": "error", "message": "公司文件夹不存在"}
            report_rows.append(row)
            append_progress(row)
            print(f"[{index}/{len(good_names)}] {company} error: 公司文件夹不存在", flush=True)
            continue

        if company_done(folder) and not args.rerun:
            data = load_existing_company_result(folder)
            if data:
                results.append({"company": company, "data": data})
                row = {
                    "company": company,
                    "status": "skipped_existing",
                    "message": "已存在正式结果，跳过",
                    "xlsx": str(folder / COMPANY_XLSX),
                    "json": str(folder / COMPANY_JSON),
                    "elapsed_seconds": round(time.time() - started, 2),
                }
                report_rows.append(row)
                append_progress(row)
                print(f"[{index}/{len(good_names)}] {company} skipped_existing", flush=True)
                continue

        try:
            mapping_info = mapping.get(company.lower(), {})
            prompt = build_prompt(company, folder, mapping_info)
            if not prompt.strip():
                raise RuntimeError("模型输入为空")
            raw = call_doubao_raw(client, prompt)
            raw_data = parse_doubao_json(raw)
            data = normalize_result(raw_data, mapping_info)
            write_company_outputs(folder, data, raw)
            results.append({"company": company, "data": data})
            row = {
                "company": company,
                "status": "success",
                "message": "",
                "xlsx": str(folder / COMPANY_XLSX),
                "json": str(folder / COMPANY_JSON),
                "elapsed_seconds": round(time.time() - started, 2),
            }
            report_rows.append(row)
            append_progress(row)
            print(f"[{index}/{len(good_names)}] {company} success", flush=True)
        except Exception as exc:
            try:
                if "raw" in locals():
                    (folder / COMPANY_RAW).write_text(raw or "", encoding="utf-8")
            except Exception:
                pass
            row = {
                "company": company,
                "status": "error",
                "message": f"{type(exc).__name__}: {exc}",
                "elapsed_seconds": round(time.time() - started, 2),
            }
            report_rows.append(row)
            append_progress(row)
            print(f"[{index}/{len(good_names)}] {company} error: {row['message']}", flush=True)

        if args.sleep > 0:
            time.sleep(args.sleep)

    write_final_workbooks(results, report_rows)
    print(TOTAL_XLSX, flush=True)
    print(REPORT_XLSX, flush=True)
    print(f"processed={len(report_rows)} success_or_existing={len(results)}", flush=True)


if __name__ == "__main__":
    main()
