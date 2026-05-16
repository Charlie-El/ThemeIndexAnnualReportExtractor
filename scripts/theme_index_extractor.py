import argparse
import hashlib
import json
import re
import traceback
import warnings
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from bs4 import BeautifulSoup, XMLParsedAsHTMLWarning
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

import annual_report_blocks as report_blocks


ROOT_DIR = Path.cwd()
DATA_DIR = ROOT_DIR / "data"
OUTPUTS_DIR = ROOT_DIR / "outputs"
DOWNLOADS_DIR = DATA_DIR / "downloads"
TOTAL_OUTPUT = OUTPUTS_DIR / "主题指数提取总表.xlsx"
RUN_REPORT = OUTPUTS_DIR / "主题指数提取运行报告.xlsx"
KEYWORD_OVERRIDES = ROOT_DIR / "templates" / "keyword_overrides.json"
PER_COMPANY_OUTPUT = "主题指数提取结果.xlsx"
HTML_SUFFIXES = {".htm", ".html"}
CELL_LIMIT = 32000
MAX_EVIDENCE_PER_GROUP = 6
MAX_THEME_EVIDENCE_PER_THEME = 4

warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)


THEME_DICTIONARY = {
    "THEME_AI": {
        "name": "AI",
        "core": [
            "artificial intelligence",
            "generative ai",
            "agentic ai",
            "machine learning",
            "deep learning",
            "large language model",
            "foundation model",
            "ai infrastructure",
            "accelerated computing",
            "data center gpu",
            "neural network",
            "computer vision",
            "natural language processing",
        ],
        "related": [
            "cloud ai",
            "model training",
            "inference",
            "automation",
            "data analytics",
            "recommendation system",
            "cuda",
            "gpu",
            "tensor",
        ],
    },
    "THEME_SEMICONDUCTOR": {
        "name": "半导体",
        "core": [
            "semiconductor",
            "integrated circuit",
            "microcontroller",
            "processor",
            "chip",
            "wafer",
            "foundry",
            "fabless",
            "memory",
            "analog",
            "logic",
            "sensor",
            "rf",
            "fpga",
            "asic",
        ],
        "related": [
            "lithography",
            "etch",
            "deposition",
            "packaging",
            "test equipment",
            "electronic design automation",
            "eda",
        ],
    },
    "THEME_ROBOTICS": {
        "name": "机器人",
        "core": [
            "robot",
            "robotics",
            "autonomous robot",
            "industrial automation",
            "factory automation",
            "machine vision",
            "motion control",
            "surgical robot",
            "autonomous vehicle",
            "self-driving",
        ],
        "related": ["cobot", "humanoid", "lidar", "actuator", "servo", "robotic automation"],
    },
    "THEME_SPACE": {
        "name": "商业航天",
        "core": [
            "space",
            "satellite",
            "launch vehicle",
            "rocket",
            "spacecraft",
            "low earth orbit",
            "leo",
            "earth observation",
            "space systems",
            "aerospace",
        ],
        "related": [
            "satellite communications",
            "ground station",
            "payload",
            "mission",
            "orbital",
            "defense aerospace",
        ],
    },
    "THEME_GEOPOLITICAL_RISK": {
        "name": "地缘风险",
        "core": [
            "defense",
            "missile",
            "military",
            "naval",
            "munition",
            "radar",
            "surveillance",
            "cybersecurity",
            "classified",
            "national security",
            "export control",
            "sanction",
        ],
        "related": ["government customer", "dod", "department of defense", "army", "navy", "air force"],
    },
    "THEME_ENERGY": {
        "name": "能源",
        "core": [
            "oil and gas",
            "natural gas",
            "lng",
            "crude oil",
            "upstream",
            "midstream",
            "refining",
            "renewable energy",
            "solar",
            "wind",
            "power generation",
            "electric utility",
        ],
        "related": ["drilling", "pipeline", "energy storage", "battery"],
    },
    "THEME_MINERALS": {
        "name": "矿产资源",
        "core": [
            "mining",
            "copper",
            "gold",
            "silver",
            "aluminum",
            "lithium",
            "nickel",
            "rare earth",
            "mineral",
            "ore",
            "smelting",
            "precious metals",
        ],
        "related": ["reserves", "mine", "royalty", "streaming", "metallurgical", "zinc", "molybdenum"],
    },
    "THEME_BIOMEDICAL": {
        "name": "生物医疗",
        "core": [
            "biotechnology",
            "biopharmaceutical",
            "clinical trial",
            "drug candidate",
            "therapeutics",
            "vaccine",
            "gene therapy",
            "cell therapy",
            "oncology",
            "medical device",
            "diagnostics",
        ],
        "related": ["fda", "phase 1", "phase 2", "phase 3", "patient", "platform technology", "antibody"],
    },
}


FIELD_GROUPS = {
    "segment": {
        "label": "分部信息",
        "priority": 1,
        "patterns": [
            r"segment information",
            r"business segments?",
            r"reportable segments?",
            r"operating segments?",
            r"operating and reportable segments?",
            r"our segments?",
            r"segment revenues?",
            r"segment net sales",
            r"segment profit",
            r"segment operating income",
            r"segment measure",
            r"geographic information",
            r"geographical information",
            r"chief operating decision maker",
            r"codm",
            r"single reportable segment",
            r"one reportable segment",
        ],
    },
    "revenue_table": {
        "label": "收入构成",
        "priority": 2,
        "patterns": [
            r"disaggregated revenues?",
            r"disaggregation of revenues?",
            r"revenue by",
            r"revenues by",
            r"net sales by",
            r"net revenue by",
            r"total net sales",
            r"revenue from",
            r"revenues from",
            r"sales from",
            r"sales by product",
            r"sales by category",
            r"sales by geography",
            r"sales by segment",
            r"product revenue",
            r"service revenue",
            r"subscription revenue",
            r"license revenue",
            r"contract revenue",
            r"revenue by geographic",
            r"revenue by region",
            r"revenue by product",
            r"revenue by service",
            r"major customers?",
            r"customer concentration",
            r"concentration of revenue",
        ],
    },
    "business": {
        "label": "业务概述",
        "priority": 3,
        "patterns": [
            r"item\s+1\.?\s+business",
            r"item\s+4\.?\s+information on the company",
            r"business overview",
            r"company overview",
            r"recent developments",
            r"strategy",
            r"our strategy",
            r"our business",
            r"our company",
            r"our businesses",
            r"description of business",
            r"overview of the business",
            r"products and services",
            r"principal products",
            r"principal services",
            r"markets and distribution",
        ],
    },
    "mda": {
        "label": "管理层讨论",
        "priority": 3,
        "patterns": [
            r"item\s+7\.?\s+management",
            r"item\s+5\.?\s+operating and financial review",
            r"management.s discussion and analysis",
            r"results of operations",
            r"operating results",
            r"financial condition",
            r"year ended",
            r"revenue for fiscal",
            r"net sales for fiscal",
        ],
    },
    "revenue_recognition": {
        "label": "收入确认",
        "priority": 4,
        "patterns": [
            r"revenue recognition",
            r"recognition of revenue",
            r"revenue from contracts with customers",
            r"contract with customers",
            r"contracts with customers",
            r"performance obligations?",
            r"transaction price",
            r"when control .* transferred",
            r"recognize revenue",
            r"recognized when",
            r"deferred revenue",
            r"contract liabilities?",
        ],
    },
}


def append_unique(target: list[str], values: list[str]) -> None:
    seen = {item.lower() for item in target}
    for value in values:
        value = str(value or "").strip()
        if not value or value.lower() in seen:
            continue
        target.append(value)
        seen.add(value.lower())


def load_keyword_overrides(path: Path = KEYWORD_OVERRIDES) -> None:
    if not path.exists():
        return
    data = json.loads(path.read_text(encoding="utf-8"))
    for group, patterns in data.get("field_groups", {}).items():
        if group in FIELD_GROUPS and isinstance(patterns, list):
            append_unique(FIELD_GROUPS[group]["patterns"], patterns)
    for keyword in data.get("table_keywords", []):
        append_unique(TABLE_KEYWORDS, [keyword])
    for theme_id, theme_cfg in data.get("theme_dictionary", {}).items():
        if theme_id not in THEME_DICTIONARY:
            THEME_DICTIONARY[theme_id] = {"name": theme_cfg.get("name", theme_id), "core": [], "related": []}
        append_unique(THEME_DICTIONARY[theme_id].setdefault("core", []), theme_cfg.get("core", []))
        append_unique(THEME_DICTIONARY[theme_id].setdefault("related", []), theme_cfg.get("related", []))

TABLE_KEYWORDS = [
    "revenue",
    "revenues",
    "net sales",
    "net revenue",
    "sales",
    "segment",
    "segments",
    "product",
    "products",
    "service",
    "services",
    "geographic",
    "geographical",
    "region",
    "regions",
    "customer",
    "customers",
    "gross profit",
    "gross margin",
    "operating income",
    "subscription",
    "license",
    "data center",
    "gaming",
    "automotive",
    "americas",
    "europe",
    "china",
    "asia",
]

SECTION_DEFINITIONS = {
    "business": {
        "label": "业务概述",
        "start": [
            r"\bitem\s+1\.?\s+business\b",
            r"\bitem\s+4\.?\s+information\s+on\s+the\s+company\b",
            r"\bbusiness\s+overview\b",
            r"\boverview\s+of\s+the\s+business\b",
            r"\bour\s+company\b",
            r"\bour\s+business(?:es)?\b",
            r"\bdescription\s+of\s+business\b",
        ],
        "end": [
            r"\bitem\s+1a\.?\s+risk\s+factors\b",
            r"\bitem\s+4a\.?\s+unresolved\s+staff\s+comments\b",
            r"\bitem\s+5\.?\s+operating\s+and\s+financial\s+review\b",
            r"\brisk\s+factors\b",
        ],
    },
    "mda": {
        "label": "管理层讨论",
        "start": [
            r"\bitem\s+7\.?\s+management.?s\s+discussion\s+and\s+analysis\b",
            r"\bitem\s+5\.?\s+operating\s+and\s+financial\s+review\b",
            r"\bmanagement.?s\s+discussion\s+and\s+analysis\b",
            r"\bresults\s+of\s+operations\b",
            r"\boperating\s+results\b",
        ],
        "end": [
            r"\bitem\s+7a\.?\s+quantitative\s+and\s+qualitative\b",
            r"\bitem\s+8\.?\s+financial\s+statements\b",
            r"\bitem\s+6\.?\s+directors\b",
            r"\bquantitative\s+and\s+qualitative\s+disclosures\b",
        ],
    },
    "financial_notes": {
        "label": "财务报表及附注",
        "start": [
            r"\bitem\s+8\.?\s+financial\s+statements\b",
            r"\bitem\s+18\.?\s+financial\s+statements\b",
            r"\bexhibits\s+financial\s+statements\b",
            r"\bnotes\s+to\s+(?:the\s+)?consolidated\s+financial\s+statements\b",
            r"\bnotes\s+to\s+(?:the\s+)?financial\s+statements\b",
        ],
        "end": [
            r"\bitem\s+9\.?\s+changes\s+in\s+and\s+disagreements\b",
            r"\bitem\s+19\.?\s+exhibits\b",
            r"\bsignatures\b",
        ],
    },
}


@dataclass
class Evidence:
    group: str
    label: str
    priority: int
    keyword: str
    position: int
    excerpt: str
    score: int = 0


@dataclass
class TableCandidate:
    table_id: int
    score: int
    keywords: str
    text: str
    latest_year: int = 0
    source_section: str = ""
    heading: str = ""


@dataclass
class ThemeHit:
    theme_id: str
    theme_name: str
    keyword: str
    keyword_type: str
    position: int
    excerpt: str
    score: int


def safe_cell(value: str, limit: int = CELL_LIMIT) -> str:
    value = "" if value is None else str(value)
    if len(value) <= limit:
        return value
    return value[: limit - 30] + "\n...[已截断]"


def clean_space(value: str) -> str:
    value = re.sub(r"\u200b", "", value or "")
    value = re.sub(r"[ \t\r\f\v]+", " ", value)
    value = re.sub(r"\n\s*\n+", "\n", value)
    return value.strip()


def normalize_line(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def infer_source_section(position: int, evidence: list[Evidence]) -> str:
    section_candidates = [item for item in evidence if item.priority == 0 and item.position <= position]
    if not section_candidates:
        return ""
    section = sorted(section_candidates, key=lambda item: item.position)[-1]
    if section.group == "financial_notes":
        return "notes"
    return section.group


def parse_report_meta(path: Path) -> tuple[str, str]:
    form_match = re.search(r"Form\s+([A-Za-z0-9-]+)", path.name, re.I)
    dates = re.findall(r"20\d{2}-\d{2}-\d{2}", path.name)
    return (form_match.group(1).upper() if form_match else "", dates[-1] if dates else "")


def html_files(folder: Path) -> list[Path]:
    return sorted(
        [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in HTML_SUFFIXES],
        key=lambda p: p.name.lower(),
    )


def remove_noise(soup: BeautifulSoup) -> None:
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()

    for tag in soup.find_all(True):
        if getattr(tag, "attrs", None) is None:
            continue
        name = (tag.name or "").lower()
        style = (tag.get("style") or "").lower()
        if name.startswith("ix:") or name in {"ix:header", "ix:hidden", "ix:resources"}:
            tag.decompose()
            continue
        if "display:none" in style.replace(" ", "") or "visibility:hidden" in style.replace(" ", ""):
            tag.decompose()


def parse_html(path: Path) -> tuple[str, BeautifulSoup]:
    return report_blocks.read_html_document(path)


def compile_patterns(patterns: list[str]) -> list[re.Pattern]:
    return [re.compile(pattern, re.I) for pattern in patterns]


def add_excerpt(
    text: str,
    group: str,
    label: str,
    priority: int,
    keyword: str,
    position: int,
    radius: int,
) -> Evidence:
    start = max(0, position - radius // 3)
    end = min(len(text), position + radius)
    excerpt = clean_space(text[start:end])
    return Evidence(group, label, priority, keyword, position, excerpt)


def add_priority_blocks(text: str, evidence: list[Evidence]) -> list[Evidence]:
    sections = report_blocks.extract_sections(text)
    blocks = report_blocks.extract_priority_blocks(text, sections)
    seen = {
        hashlib.md5(item.excerpt[:1000].encode("utf-8", errors="ignore")).hexdigest()
        for item in evidence
    }
    for block in blocks:
        if block.source_section == "notes":
            group = "segment" if "segment" in block.label.lower() else "revenue_table"
        elif block.source_section == "business":
            group = "business"
        elif block.source_section == "mda":
            group = "mda" if "revenue recognition" not in block.label.lower() else "revenue_recognition"
        else:
            group = "revenue_table"
        item = Evidence(
            group=group,
            label=block.label,
            priority=block.priority,
            keyword=block.matched_term,
            position=block.start,
            excerpt=block.text,
            score=110 - block.priority,
        )
        digest = hashlib.md5(item.excerpt[:1000].encode("utf-8", errors="ignore")).hexdigest()
        if digest in seen:
            continue
        seen.add(digest)
        evidence.append(item)
    return evidence


def looks_like_table_of_contents(block: str) -> bool:
    compact = normalize_line(block[:1200]).lower()
    item_count = len(re.findall(r"\bitem\s+\d+[a-z]?\.", compact, re.I))
    dotted_count = compact.count("....")
    return item_count >= 4 or dotted_count >= 2 or "table of contents" in compact[:200]


def find_nearest_end(text: str, start: int, end_patterns: list[str]) -> int | None:
    nearest = None
    for pattern in end_patterns:
        for match in re.finditer(pattern, text[start + 50 :], re.I):
            candidate = start + 50 + match.start()
            if candidate > start:
                nearest = candidate if nearest is None else min(nearest, candidate)
                break
    return nearest


def find_section_bounds(text: str, start_patterns: list[str], end_patterns: list[str]) -> tuple[int | None, int | None, str]:
    start_hits = []
    for pattern in start_patterns:
        for match in re.finditer(pattern, text, re.I):
            if match.start() < 1500:
                continue
            start_hits.append((match.start(), match.group(0)))
    if not start_hits:
        return None, None, ""

    candidates = []
    for start, keyword in sorted(start_hits, key=lambda item: item[0]):
        preview = text[start : start + 1500]
        if looks_like_table_of_contents(preview):
            continue
        nearest_end = find_nearest_end(text, start, end_patterns)
        section_length = (nearest_end - start) if nearest_end else 60000
        if section_length < 1800:
            continue
        candidates.append((start, nearest_end, keyword, section_length))

    if not candidates:
        start, keyword = sorted(start_hits, key=lambda item: item[0])[-1]
        nearest_end = find_nearest_end(text, start, end_patterns)
        return start, min(nearest_end or len(text), start + 60000), keyword

    start, nearest_end, keyword, _ = sorted(candidates, key=lambda item: item[0])[0]
    return start, min(nearest_end or len(text), start + 60000), keyword


def find_section_evidence(text: str) -> list[Evidence]:
    evidence = []
    structured_sections = report_blocks.extract_sections(text)
    for section in structured_sections:
        group = "financial_notes" if section.source_section == "notes" else section.source_section
        excerpt = clean_space(text[section.start : min(section.end, section.start + 6500)])
        evidence.append(
            Evidence(
                group=group,
                label=section.name,
                priority=0,
                keyword=section.matched_heading,
                position=section.start,
                excerpt=excerpt,
                score=100,
            )
        )
    if evidence:
        return evidence
    for group, cfg in SECTION_DEFINITIONS.items():
        start, end, keyword = find_section_bounds(text, cfg["start"], cfg["end"])
        if start is None or end is None:
            continue
        excerpt = clean_space(text[start:min(end, start + 6500)])
        evidence.append(
            Evidence(
                group=group,
                label=cfg["label"],
                priority=0,
                keyword=keyword,
                position=start,
                excerpt=excerpt,
                score=100,
            )
        )
    return evidence


def score_evidence(group: str, excerpt: str) -> int:
    lower = excerpt.lower()
    score = 0
    positive_terms = {
        "segment": [
            "revenue",
            "net sales",
            "operating income",
            "reportable",
            "chief operating decision maker",
            "codm",
            "product",
            "geographic",
            "gross margin",
            "segment profit",
        ],
        "revenue_table": [
            "revenue",
            "revenues",
            "net sales",
            "customer",
            "contract",
            "product",
            "service",
            "year ended",
            "total",
            "geographic",
            "region",
            "subscription",
            "license",
        ],
        "revenue_recognition": [
            "recognize",
            "recognized",
            "performance obligation",
            "contract",
            "customer",
            "transaction price",
            "control",
            "deferred revenue",
        ],
        "business": [
            "we design",
            "we develop",
            "we manufacture",
            "we provide",
            "we offer",
            "products",
            "services",
            "customers",
            "business",
            "platform",
            "solutions",
        ],
        "mda": ["results of operations", "net sales", "revenue", "gross margin", "operating income", "year ended"],
    }
    for term in positive_terms.get(group, []):
        if term in lower:
            score += 8
    if re.search(r"\b20\d{2}\b", excerpt):
        score += 4
    if re.search(r"\d", excerpt):
        score += 2
    negative_terms = ["risk factor", "forward-looking", "safe harbor", "table of contents", "cover page"]
    for term in negative_terms:
        if term in lower:
            score -= 12
    if looks_like_table_of_contents(excerpt):
        score -= 30
    return score


def find_theme_hits(text: str, evidence: list[Evidence]) -> list[ThemeHit]:
    hits: list[ThemeHit] = []
    seen = set()
    for theme_id, cfg in THEME_DICTIONARY.items():
        for keyword_type in ("core", "related"):
            for keyword in cfg[keyword_type]:
                pattern = re.compile(r"(?<![A-Za-z0-9])" + re.escape(keyword) + r"(?![A-Za-z0-9])", re.I)
                theme_hits = []
                for match in pattern.finditer(text):
                    key = (theme_id, keyword.lower(), match.start() // 2500)
                    if key in seen:
                        continue
                    seen.add(key)
                    excerpt = clean_space(text[max(0, match.start() - 900) : min(len(text), match.start() + 1600)])
                    score = 10 if keyword_type == "core" else 5
                    if infer_source_section(match.start(), evidence) in {"business", "mda", "notes", "financial_notes"}:
                        score += 4
                    if any(term in excerpt.lower() for term in ["revenue", "net sales", "segment", "customer", "product", "service"]):
                        score += 3
                    if looks_like_table_of_contents(excerpt):
                        score -= 10
                    theme_hits.append(
                        ThemeHit(
                            theme_id=theme_id,
                            theme_name=cfg["name"],
                            keyword=keyword,
                            keyword_type=keyword_type,
                            position=match.start(),
                            excerpt=excerpt,
                            score=score,
                        )
                    )
                theme_hits.sort(key=lambda item: (-item.score, item.position))
                hits.extend(theme_hits[:MAX_THEME_EVIDENCE_PER_THEME])
    hits.sort(key=lambda item: (item.theme_id, -item.score, item.position))
    return hits


def find_evidence(text: str) -> list[Evidence]:
    evidence: list[Evidence] = find_section_evidence(text)
    evidence = add_priority_blocks(text, evidence)
    seen_hashes = set()
    for item in evidence:
        seen_hashes.add(hashlib.md5(item.excerpt[:1000].encode("utf-8", errors="ignore")).hexdigest())
    financial_positions = [item.position for item in evidence if item.group in {"financial_notes", "mda"}]
    for group, cfg in FIELD_GROUPS.items():
        hits = []
        for pattern in compile_patterns(cfg["patterns"]):
            for match in pattern.finditer(text):
                hits.append((match.start(), match.group(0)))
        hits.sort(key=lambda item: item[0])

        selected = []
        if group in {"segment", "revenue_table", "revenue_recognition"} and financial_positions:
            min_financial = min(financial_positions)
            hits = [hit for hit in hits if hit[0] >= min_financial - 5000] or hits

        scored_hits = []
        for position, keyword in hits:
            preview = clean_space(text[max(0, position - 500): min(len(text), position + 2600)])
            preview_score = score_evidence(group, preview)
            if group in {"segment", "revenue_table", "revenue_recognition"}:
                if financial_positions:
                    distance = min(abs(position - pos) for pos in financial_positions)
                    preview_score += max(0, 20 - distance // 10000)
                if position < 8000:
                    preview_score -= 20
            scored_hits.append((preview_score, position, keyword))
        scored_hits.sort(key=lambda item: (-item[0], item[1]))

        for _, position, keyword in scored_hits:
            if any(abs(position - old_position) < 1800 for old_position, _ in selected):
                continue
            selected.append((position, keyword))
            if len(selected) >= 5:
                break

        for position, keyword in selected:
            item = add_excerpt(
                text=text,
                group=group,
                label=cfg["label"],
                priority=cfg["priority"],
                keyword=keyword,
                position=position,
                radius=4200,
            )
            item.score = score_evidence(group, item.excerpt)
            digest = hashlib.md5(item.excerpt[:1000].encode("utf-8", errors="ignore")).hexdigest()
            if digest in seen_hashes:
                continue
            seen_hashes.add(digest)
            evidence.append(item)

    evidence.sort(key=lambda item: (item.priority, item.group, -item.score, item.position))
    return evidence


def table_to_text(table) -> str:
    rows = []
    for tr in table.find_all("tr"):
        cells = []
        for cell in tr.find_all(["th", "td"]):
            cell_text = normalize_line(cell.get_text(" "))
            if cell_text:
                cells.append(cell_text)
        if cells:
            rows.append(" | ".join(cells))
    return "\n".join(rows)


def score_table(text: str) -> tuple[int, list[str]]:
    lower = text.lower()
    matched = [kw for kw in TABLE_KEYWORDS if kw in lower]
    if not matched:
        return 0, []

    score = 0
    for kw in matched:
        if kw in {"revenue", "revenues", "net sales", "net revenue"}:
            score += 12
        elif kw in {"segment", "segments"}:
            score += 12
        elif kw in {"product", "products", "service", "services", "geographic", "geographical", "region", "regions"}:
            score += 5
        elif kw in {"gross profit", "gross margin", "operating income"}:
            score += 4
        else:
            score += 2
    if re.search(r"\b20\d{2}\b", text):
        score += 4
    if re.search(r"\d", text):
        score += 3
    if "total" in lower:
        score += 2
    if ("revenue" in lower or "net sales" in lower or "sales" in lower) and any(
        term in lower
        for term in [
            "segment",
            "product",
            "service",
            "geographic",
            "region",
            "customer",
            "subscription",
            "license",
            "data center",
            "gaming",
            "automotive",
        ]
    ):
        score += 18
    if re.search(r"\bitem\s+\d+[a-z]?\.", lower) and ("part i" in lower or "part ii" in lower):
        score -= 45
    if any(term in lower for term in ["assets:", "liabilities", "stockholders", "cash and cash equivalents"]):
        score -= 12
    return score, matched


def find_table_candidates(soup: BeautifulSoup) -> list[TableCandidate]:
    text = clean_space(soup.get_text("\n"))
    sections = report_blocks.extract_sections(text)
    structured_tables = report_blocks.extract_table_blocks(soup, text, sections, limit=18)
    if structured_tables:
        return [
            TableCandidate(
                table_id=item.table_id,
                score=item.score,
                keywords=item.keywords,
                text=item.text,
                latest_year=item.latest_year,
                source_section=item.source_section,
                heading=item.heading,
            )
            for item in structured_tables
        ]

    candidates: list[TableCandidate] = []
    for index, table in enumerate(soup.find_all("table"), start=1):
        table_text = table_to_text(table)
        if len(table_text) < 40:
            continue
        score, keywords = score_table(table_text)
        if score <= 0:
            continue
        candidates.append(
            TableCandidate(
                table_id=index,
                score=score,
                keywords=", ".join(keywords),
                text=clean_space(table_text),
                latest_year=report_blocks.latest_year(table_text),
            )
        )
    candidates.sort(key=lambda item: (item.score, item.latest_year, len(item.text)), reverse=True)
    return candidates[:18]


def best_excerpt(evidence: list[Evidence], group: str) -> str:
    group_items = [item for item in evidence if item.group == group]
    if not group_items:
        return ""
    preferred = [item for item in group_items if item.priority == 0]
    if preferred:
        return safe_cell(preferred[0].excerpt, 3000)
    non_toc = [item for item in group_items if not looks_like_table_of_contents(item.excerpt)]
    ranked = sorted(non_toc or group_items, key=lambda item: (item.score, -item.position), reverse=True)
    return safe_cell(ranked[0].excerpt, 3000)


def best_table_text(tables: list[TableCandidate]) -> str:
    if not tables:
        return ""
    return safe_cell(tables[0].text, 5000)


def disclosure_dims(evidence: list[Evidence], tables: list[TableCandidate]) -> list[str]:
    dims = set()
    combined = "\n".join([item.excerpt for item in evidence] + [item.text for item in tables]).lower()
    if any(term in combined for term in ["reportable segment", "operating segment", "business segment", "segment revenue"]):
        dims.add("business_segment")
    if any(term in combined for term in ["product", "products", "service", "services", "subscription", "license"]):
        dims.add("product")
    if any(term in combined for term in ["industry", "industries", "end market", "vertical"]):
        dims.add("industry")
    if any(term in combined for term in ["geographic", "geographical", "region", "americas", "europe", "china", "asia"]):
        dims.add("region")
    if any(term in combined for term in ["customer", "customers", "direct customer", "indirect customer"]):
        dims.add("customer_type")
    if not dims and best_excerpt(evidence, "business"):
        dims.add("narrative_only")
    return sorted(dims)


def infer_dim_from_text(value: str) -> str:
    lower = (value or "").lower()
    if any(
        term in lower
        for term in [
            "reportable segment",
            "operating segment",
            "business segment",
            "segment revenue",
            "revenue by segment",
            "segment adjusted gross margin",
        ]
    ) or ("segment" in lower and ("revenue" in lower or "net sales" in lower)):
        return "business_segment"
    if any(term in lower for term in ["product", "products", "service", "services", "magnetic sensors", "power integrated circuits", "contract type"]):
        return "product"
    if any(term in lower for term in ["industry", "industries", "end market", "vertical", "automotive", "industrial"]):
        return "industry"
    if any(term in lower for term in ["geographic", "geographical", "region", "domestic", "international", "americas", "europe", "china", "asia"]):
        return "region"
    if any(term in lower for term in ["customer", "customers", "u.s. government", "non-u.s. government"]):
        return "customer_type"
    return "narrative_only" if value else ""


def dimension_scores(evidence: list[Evidence], tables: list[TableCandidate]) -> dict[str, int]:
    scores: dict[str, int] = {}
    for table in tables:
        dim = infer_dim_from_text("\n".join([table.heading, table.keywords, table.text]))
        if not dim:
            continue
        score = table.score + (30 if table.latest_year else 0)
        if "revenue" in table.text.lower() or "net sales" in table.text.lower():
            score += 25
        if re.search(r"\d", table.text):
            score += 10
        scores[dim] = max(scores.get(dim, 0), score)
    for item in evidence:
        dim = infer_dim_from_text(item.excerpt)
        if not dim:
            continue
        score = item.score + max(0, 40 - item.priority * 8)
        if "revenue" in item.excerpt.lower() or "net sales" in item.excerpt.lower():
            score += 20
        scores[dim] = max(scores.get(dim, 0), score)
    return scores


def primary_disclosure_dim(dims: list[str], evidence: list[Evidence] | None = None, tables: list[TableCandidate] | None = None) -> str:
    if evidence is not None or tables is not None:
        scores = dimension_scores(evidence or [], tables or [])
        if scores:
            return sorted(scores.items(), key=lambda item: (-item[1], item[0]))[0][0]
    for dim in ["business_segment", "product", "industry", "region", "customer_type", "narrative_only"]:
        if dim in dims:
            return dim
    return ""


def data_quality_flag(evidence: list[Evidence], tables: list[TableCandidate], quality_note: str = "") -> tuple[str, float]:
    if "NT延迟提交通知" in quality_note or "未包含财务报表" in quality_note:
        return "E（材料不足）", 0.0
    dims = disclosure_dims(evidence, tables)
    table_text = "\n".join(item.text for item in tables).lower()
    evidence_text = "\n".join(item.excerpt for item in evidence).lower()
    has_segment = "business_segment" in dims or "segment" in evidence_text
    revenue_tables = [
        table for table in tables
        if re.search(r"\brevenue[s]?\b|\bnet sales\b|\bproduct sales\b|\bcontract services\b", table.text, re.I)
    ]
    has_revenue = bool(revenue_tables) or "revenue" in evidence_text or "net sales" in evidence_text
    has_ratio = "%" in table_text or "percentage" in table_text or "ratio" in table_text
    has_product_or_industry = bool({"product", "industry", "region", "customer_type"} & set(dims))
    if has_segment and has_revenue and (has_ratio or tables):
        return "A（分部+营收+占比）", 1.0
    if has_product_or_industry and has_revenue:
        return "B（产品/行业+营收）", 0.85
    if has_revenue:
        return "C（有营收无分解）", 0.65
    if best_excerpt(evidence, "business"):
        return "D（仅文字描述）", 0.4
    return "E（材料不足）", 0.0


def has_revenue_breakdown(evidence: list[Evidence], tables: list[TableCandidate]) -> bool:
    if any(re.search(r"\brevenue[s]?\b|\bnet sales\b|\bproduct sales\b|\bcontract services\b", table.text, re.I) for table in tables):
        return True
    return any(item.group == "revenue_table" and re.search(r"\brevenue[s]?\b|\bnet sales\b", item.excerpt, re.I) for item in evidence)


def adjust_status_for_quality(status: str, note: str, quality_note: str) -> tuple[str, str]:
    if not quality_note:
        return status, note
    risk_note = quality_note
    if note:
        risk_note = f"{note}；{quality_note}"
    if "NT延迟提交通知" in quality_note or "未包含财务报表" in quality_note:
        return "未找到候选内容", risk_note
    if "疑似年报修订版" in quality_note and status == "已提取候选内容":
        return "部分提取", risk_note
    return status, risk_note


def theme_summary(theme_hits: list[ThemeHit]) -> str:
    if not theme_hits:
        return ""
    grouped = {}
    for hit in theme_hits:
        grouped.setdefault(hit.theme_id, []).append(hit)
    rows = []
    for theme_id, hits in grouped.items():
        hits = sorted(hits, key=lambda item: (-item.score, item.position))
        keywords = []
        for hit in hits:
            if hit.keyword not in keywords:
                keywords.append(hit.keyword)
            if len(keywords) >= 8:
                break
        score = min(1.0, sum(max(hit.score, 0) for hit in hits[:8]) / 80)
        rows.append(f"{hits[0].theme_name}({theme_id}): score={score:.2f}; keywords={', '.join(keywords)}")
    return safe_cell("\n".join(rows), 5000)


def feedback_instruction(company: str) -> str:
    payload = {
        "purpose": "如果当前证据不足，请只提出关键词/章节定位建议，不要修改代码。",
        "company_name": company,
        "return_json_schema": {
            "missing_fields": ["main_business_desc", "segments", "revenue_breakdown", "revenue_recognition", "theme_mapping"],
            "suggested_keywords": [
                {
                    "field_group": "business/segment/revenue_table/revenue_recognition/theme",
                    "keyword_or_regex": "建议新增的英文关键词或正则",
                    "keyword_type": "core/related/excluded",
                    "reason_cn": "为什么这个关键词可以提升召回",
                    "evidence_quote": "年报原文短句，不超过60词",
                    "source_section": "business/MDA/notes/segment/table/unknown",
                }
            ],
            "suspected_missing_sections": ["可能缺失的章节名"],
            "do_not_modify_code": True,
        },
    }
    return json.dumps(payload, ensure_ascii=False, indent=2)


def combined_context(
    company: str,
    report_file: str,
    evidence: list[Evidence],
    tables: list[TableCandidate],
    theme_hits: list[ThemeHit],
    quality_note: str = "",
) -> str:
    dims = disclosure_dims(evidence, tables)
    flag, confidence = data_quality_flag(evidence, tables, quality_note)
    payload = {
        "task": "请根据美股年报证据，按主题指数数据维护字段提取主营业务、分部明细、收入构成、收入确认、主题相关性映射。只输出合法JSON，不要编造。",
        "company": company,
        "report_file": report_file,
        "reading_order": [
            "优先使用 PDF 需求优先证据：P1 Item 8/Notes Segment Information；P2 Item 8/Notes REVENUES；P3 Item 1 Business 或 Item 7 MD&A；P4 Revenue Recognition。",
            "同一字段有多个候选时，优先使用最新 fiscal year 的完整表格；较低优先级来源如果更完整，可补充但必须保留来源。",
            "evidence 必须来自下方原文或表格行，不要改写；没有证据的字段留空。",
        ],
        "required_output_fields": [
            "report_period",
            "report_type",
            "主营业务描述",
            "主营业务标准化摘要",
            "业务/经营分部",
            "收入构成",
            "毛利/毛利率构成",
            "收入确认方式",
            "主要披露维度",
            "可用披露维度",
            "主题相关性映射",
            "数据质量等级",
            "证据引用",
        ],
        "disclosure_dims_detected": dims,
        "primary_disclosure_dim": primary_disclosure_dim(dims, evidence, tables),
        "source_quality_note": quality_note,
        "data_quality_flag": flag,
        "confidence": confidence,
        "evidence": [
            {
                "type": item.label,
                "priority": item.priority,
                "keyword": item.keyword,
                "source_section": item.group,
                "excerpt": safe_cell(item.excerpt, 3200),
            }
            for item in evidence[:14]
        ],
        "tables": [
            {
                "table_id": item.table_id,
                "score": item.score,
                "latest_year": item.latest_year,
                "source_section": item.source_section,
                "heading": item.heading,
                "keywords": item.keywords,
                "text": safe_cell(item.text, 4000),
            }
            for item in tables[:8]
        ],
        "theme_hits": [
            {
                "theme_id": item.theme_id,
                "theme_name": item.theme_name,
                "keyword": item.keyword,
                "keyword_type": item.keyword_type,
                "score": item.score,
                "evidence": safe_cell(item.excerpt, 900),
            }
            for item in theme_hits[:20]
        ],
        "model_feedback_instruction": json.loads(feedback_instruction(company)),
    }
    return safe_cell(json.dumps(payload, ensure_ascii=False, indent=2), 30000)


def extraction_status(evidence: list[Evidence], tables: list[TableCandidate]) -> tuple[str, str]:
    groups = {item.group for item in evidence}
    has_core = bool({"segment", "revenue_table", "business"} & groups) or bool(tables)
    has_segment_or_revenue = bool({"segment", "revenue_table"} & groups) or bool(tables)
    if has_core and has_segment_or_revenue:
        return "已提取候选内容", ""
    if evidence:
        return "部分提取", "未稳定识别到分部/收入构成，建议人工或模型复核"
    return "未找到候选内容", "未命中业务、分部、收入相关关键词，建议人工检查年报或附件"


def source_quality_notes(report_path: Path, form_type: str, text: str) -> str:
    notes = []
    lower_name = report_path.name.lower()
    lower_text = text[:25000].lower()
    if form_type.startswith("NT") or "notification of late filing" in lower_name or "form 12b-25" in lower_text:
        notes.append("疑似NT延迟提交通知，可能不是完整年报")
    if "amend" in lower_name or "amendment no." in lower_text or "explanatory note" in lower_text:
        notes.append("疑似年报修订版，需确认是否包含业务与财务报表")
    if "no financial statements" in lower_text or "no financial statements have been included" in lower_text:
        notes.append("文件明确说明未包含财务报表")
    if form_type == "40-F" and "principal documents" in lower_text and "incorporated by reference" in lower_text:
        notes.append("40-F主要是Principal Documents/Exhibit引用页，需确认是否包含Exhibit正文")
    if len(text) < 12000:
        notes.append("HTML可读正文较短")
    return "；".join(notes)


def write_company_workbook(
    folder: Path,
    company: str,
    report_path: Path,
    form_type: str,
    report_date: str,
    status: str,
    note: str,
    evidence: list[Evidence],
    tables: list[TableCandidate],
    theme_hits: list[ThemeHit],
    quality_note: str,
) -> None:
    dims = disclosure_dims(evidence, tables)
    quality_flag, confidence = data_quality_flag(evidence, tables, quality_note)
    wb = Workbook()
    ws = wb.active
    ws.title = "提取摘要"
    summary_rows = [
        ["字段", "内容"],
        ["公司名称", company],
        ["年报类型", form_type],
        ["年报日期", report_date],
        ["年报文件", report_path.name],
        ["提取状态", status],
        ["备注", note],
        ["源文件质量提示", quality_note],
        ["主要披露维度", primary_disclosure_dim(dims, evidence, tables)],
        ["可用披露维度", json.dumps(dims, ensure_ascii=False)],
        ["是否有业务分部披露", "是" if "business_segment" in dims else "否"],
        ["是否有任何形式营收分解", "是" if has_revenue_breakdown(evidence, tables) else "否"],
        ["是否有毛利/毛利率分解", "是" if any(term in ("\n".join([item.excerpt for item in evidence] + [item.text for item in tables]).lower()) for term in ["gross profit", "gross margin"]) else "否"],
        ["数据质量等级", quality_flag],
        ["置信度", confidence],
        ["主题命中摘要", theme_summary(theme_hits)],
        ["业务概述候选", best_excerpt(evidence, "business")],
        ["管理层讨论候选", best_excerpt(evidence, "mda")],
        ["财务附注候选", best_excerpt(evidence, "financial_notes")],
        ["分部信息候选", best_excerpt(evidence, "segment")],
        ["收入构成候选", best_excerpt(evidence, "revenue_table")],
        ["收入确认候选", best_excerpt(evidence, "revenue_recognition")],
        ["最佳收入/分部表格", best_table_text(tables)],
    ]
    for row in summary_rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 110

    ev_ws = wb.create_sheet("候选证据")
    ev_ws.append(["优先级", "类型", "关键词", "位置", "得分", "证据片段"])
    for item in evidence:
        ev_ws.append([item.priority, item.label, item.keyword, item.position, item.score, safe_cell(item.excerpt)])
    ev_ws.column_dimensions["A"].width = 10
    ev_ws.column_dimensions["B"].width = 18
    ev_ws.column_dimensions["C"].width = 30
    ev_ws.column_dimensions["D"].width = 12
    ev_ws.column_dimensions["E"].width = 10
    ev_ws.column_dimensions["F"].width = 120

    table_ws = wb.create_sheet("表格候选")
    table_ws.append(["表格序号", "得分", "最新年份", "来源章节", "附近标题", "命中关键词", "表格文本"])
    for item in tables:
        table_ws.append([item.table_id, item.score, item.latest_year, item.source_section, item.heading, item.keywords, safe_cell(item.text)])
    table_ws.column_dimensions["A"].width = 12
    table_ws.column_dimensions["B"].width = 10
    table_ws.column_dimensions["C"].width = 12
    table_ws.column_dimensions["D"].width = 14
    table_ws.column_dimensions["E"].width = 42
    table_ws.column_dimensions["F"].width = 36
    table_ws.column_dimensions["G"].width = 120

    theme_ws = wb.create_sheet("主题命中")
    theme_ws.append(["主题ID", "主题名称", "关键词", "关键词类型", "位置", "得分", "证据片段"])
    for item in theme_hits:
        theme_ws.append([item.theme_id, item.theme_name, item.keyword, item.keyword_type, item.position, item.score, safe_cell(item.excerpt)])
    theme_ws.column_dimensions["A"].width = 22
    theme_ws.column_dimensions["B"].width = 18
    theme_ws.column_dimensions["C"].width = 28
    theme_ws.column_dimensions["D"].width = 14
    theme_ws.column_dimensions["E"].width = 12
    theme_ws.column_dimensions["F"].width = 10
    theme_ws.column_dimensions["G"].width = 120

    llm_ws = wb.create_sheet("模型输入")
    llm_ws.append(["字段", "内容"])
    llm_ws.append(["建议输入", combined_context(company, report_path.name, evidence, tables, theme_hits, quality_note)])
    llm_ws.append(["模型反馈模板", feedback_instruction(company)])
    llm_ws.column_dimensions["A"].width = 18
    llm_ws.column_dimensions["B"].width = 120

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

    wb.save(folder / PER_COMPANY_OUTPUT)
    wb.close()


def process_company(folder: Path) -> dict:
    company = folder.name
    files = html_files(folder)
    if not files:
        return {
            "公司名称": company,
            "年报类型": "",
            "年报日期": "",
            "年报文件": "",
            "提取状态": "失败",
            "备注": "公司文件夹中没有html年报",
        }

    report_path = files[0]
    form_type, report_date = parse_report_meta(report_path)
    try:
        text, soup = parse_html(report_path)
        evidence = find_evidence(text)
        tables = find_table_candidates(soup)
        theme_hits = find_theme_hits(text, evidence)
        quality_note = source_quality_notes(report_path, form_type, text)
        status, note = extraction_status(evidence, tables)
        status, note = adjust_status_for_quality(status, note, quality_note)
        dims = disclosure_dims(evidence, tables)
        quality_flag, confidence = data_quality_flag(evidence, tables, quality_note)
        write_company_workbook(folder, company, report_path, form_type, report_date, status, note, evidence, tables, theme_hits, quality_note)
        return {
            "公司名称": company,
            "年报类型": form_type,
            "年报日期": report_date,
            "年报文件": report_path.name,
            "提取状态": status,
            "备注": note,
            "源文件质量提示": quality_note,
            "主要披露维度": primary_disclosure_dim(dims, evidence, tables),
            "可用披露维度": json.dumps(dims, ensure_ascii=False),
            "是否有业务分部披露": "是" if "business_segment" in dims else "否",
            "是否有任何形式营收分解": "是" if has_revenue_breakdown(evidence, tables) else "否",
            "数据质量等级": quality_flag,
            "置信度": confidence,
            "主题命中摘要": theme_summary(theme_hits),
            "业务概述候选": best_excerpt(evidence, "business"),
            "管理层讨论候选": best_excerpt(evidence, "mda"),
            "财务附注候选": best_excerpt(evidence, "financial_notes"),
            "分部信息候选": best_excerpt(evidence, "segment"),
            "收入构成候选": best_excerpt(evidence, "revenue_table"),
            "收入确认候选": best_excerpt(evidence, "revenue_recognition"),
            "最佳收入/分部表格": best_table_text(tables),
            "表格候选数量": len(tables),
            "证据片段数量": len(evidence),
            "主题命中数量": len(theme_hits),
            "企业结果文件": str(folder / PER_COMPANY_OUTPUT),
        }
    except Exception as exc:
        return {
            "公司名称": company,
            "年报类型": form_type,
            "年报日期": report_date,
            "年报文件": report_path.name,
            "提取状态": "失败",
            "备注": f"{type(exc).__name__}: {exc}",
            "错误详情": traceback.format_exc(limit=3),
        }


def write_total_workbook(rows: list[dict], output: Path, report: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    report.parent.mkdir(parents=True, exist_ok=True)

    columns = [
        "公司名称",
        "年报类型",
        "年报日期",
        "年报文件",
        "提取状态",
        "备注",
        "源文件质量提示",
        "主要披露维度",
        "可用披露维度",
        "是否有业务分部披露",
        "是否有任何形式营收分解",
        "数据质量等级",
        "置信度",
        "主题命中摘要",
        "业务概述候选",
        "管理层讨论候选",
        "财务附注候选",
        "分部信息候选",
        "收入构成候选",
        "收入确认候选",
        "最佳收入/分部表格",
        "表格候选数量",
        "证据片段数量",
        "主题命中数量",
        "企业结果文件",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "总表"
    ws.append(columns)
    for row in rows:
        ws.append([safe_cell(row.get(col, "")) for col in columns])
    widths = {
        "A": 28,
        "B": 12,
        "C": 14,
        "D": 70,
        "E": 16,
        "F": 42,
        "G": 34,
        "H": 18,
        "I": 36,
        "J": 18,
        "K": 20,
        "L": 22,
        "M": 12,
        "N": 55,
        "O": 70,
        "P": 70,
        "Q": 70,
        "R": 70,
        "S": 70,
        "T": 70,
        "U": 90,
        "V": 14,
        "W": 14,
        "X": 14,
        "Y": 80,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    wb.save(output)
    wb.close()

    report_wb = Workbook()
    summary = report_wb.active
    summary.title = "运行报告"
    status_counts = {}
    for row in rows:
        status_counts[row.get("提取状态", "")] = status_counts.get(row.get("提取状态", ""), 0) + 1
    summary.append(["项目", "数值"])
    summary.append(["运行时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary.append(["公司数量", len(rows)])
    for status, count in sorted(status_counts.items()):
        summary.append([status, count])

    issue = report_wb.create_sheet("需复核")
    issue.append(columns)
    for row in rows:
        if row.get("提取状态") != "已提取候选内容":
            issue.append([safe_cell(row.get(col, "")) for col in columns])
    for sheet in report_wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
    report_wb.save(report)
    report_wb.close()


def main() -> None:
    load_keyword_overrides()
    parser = argparse.ArgumentParser(description="Extract theme-index candidate fields from annual report HTML files.")
    parser.add_argument("--downloads-dir", default=str(DOWNLOADS_DIR))
    parser.add_argument("--limit", type=int, default=0, help="Only process the first N company folders.")
    parser.add_argument("--companies", nargs="*", default=[], help="Only process the named company folders.")
    args = parser.parse_args()

    downloads_dir = Path(args.downloads_dir)
    folders = sorted([p for p in downloads_dir.iterdir() if p.is_dir()], key=lambda p: p.name.lower())
    if args.companies:
        wanted = {name.lower() for name in args.companies}
        folders = [folder for folder in folders if folder.name.lower() in wanted]
    if args.limit > 0:
        folders = folders[: args.limit]

    rows = []
    total = len(folders)
    for index, folder in enumerate(folders, start=1):
        print(f"[{index}/{total}] {folder.name}", flush=True)
        rows.append(process_company(folder))

    write_total_workbook(rows, TOTAL_OUTPUT, RUN_REPORT)
    print(TOTAL_OUTPUT)
    print(RUN_REPORT)
    print(f"processed={len(rows)}")


if __name__ == "__main__":
    main()
